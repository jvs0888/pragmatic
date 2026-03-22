import asyncio
import re
from urllib.parse import parse_qs
from dataclasses import dataclass
import typing as t
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from loggers.logger import logger
except ImportError as ie:
    exit(f"{ie} :: {Path(__file__).resolve()}")


STARTING_BALANCE = 100_000.0


@dataclass
class SpinRecord:
    spin_number: int
    index: int
    bet: float
    win: float
    balance: float
    ntp: float
    reel_set: int
    has_tumble: bool
    is_winning: bool
    event: str = ""
    free_spins: str = ""
    multiplier: int = 0
    fs_win: float = 0.0


class SpinCollector:
    def __init__(self, spin_limit: int = 100) -> None:
        self.spin_limit: int = spin_limit
        self.raw_responses: list[str] = []
        self.spins: list[SpinRecord] = []
        self._spin_count: int = 0
        self._done: asyncio.Event = asyncio.Event()
        self._pending_spin: t.Optional[SpinRecord] = None
        self._in_free_spins: bool = False

    def is_main_spin(self, params: dict) -> bool:
        return params.get("na") == ["s"] and "s" in params

    def parse_response(self, text: str) -> t.Optional[SpinRecord]:
        params: dict = parse_qs(text)

        if not self.is_main_spin(params):
            return None

        def get(key: str, default: t.Any = None) -> t.Any:
            val: list = params.get(key, [default])
            return val[0] if val else default

        self._spin_count += 1
        bet: float = float(get("c", 0)) * float(get("l", 0))
        win: float = float(get("tw", 0))
        balance_str: str = get("balance", "0").replace(",", "")

        return SpinRecord(
            spin_number=self._spin_count,
            index=int(get("index", 0)),
            bet=bet,
            win=win,
            balance=float(balance_str),
            ntp=float(get("ntp", 0)),
            reel_set=int(get("reel_set", 0)),
            has_tumble="tmb" in params,
            is_winning=win > 0,
        )

    def process_response(self, text: str) -> None:
        self.raw_responses.append(text)
        params: dict = parse_qs(text)
        na: str = params.get("na", [""])[0]

        def get(key: str, default: t.Any = None) -> t.Any:
            val: list = params.get(key, [default])
            return val[0] if val else default

        if na == "s" and "s" in params:
            tw_param: t.Optional[str] = get("tw")
            if tw_param is None:
                return

            fs: t.Optional[str] = get("fs")

            rs_p: t.Optional[str] = get("rs_p")
            is_cascade: bool = rs_p is not None and int(rs_p) > 0

            if fs is not None:
                if fs == "1":
                    if self._pending_spin is not None:
                        self._flush_pending_spin()

                    self._in_free_spins = True
                    self._spin_count += 1
                    bet: float = float(get("c", 0)) * float(get("l", 0))

                    self._pending_spin = SpinRecord(
                        spin_number=self._spin_count,
                        index=int(get("index", 0)),
                        bet=bet,
                        win=float(get("tw", 0)),
                        balance=float(get("balance", "0").replace(",", "")),
                        ntp=float(get("ntp", 0)),
                        reel_set=int(get("reel_set", 0)),
                        has_tumble="tmb" in params,
                        is_winning=True,
                        multiplier=0,
                        event="Free Spins",
                        free_spins=get("fsmax", ""),
                        fs_win=float(get("fswin", 0)),
                    )

                    if self._spin_count >= self.spin_limit:
                        self._done.set()
                else:
                    if self._pending_spin is not None:
                        fswin: float = float(get("fswin", 0))
                        self._pending_spin.fs_win = max(self._pending_spin.fs_win, fswin)
                        self._pending_spin.win = max(self._pending_spin.win, float(get("tw", 0)))

                        balance_str: str = get("balance", "0").replace(",", "")
                        self._pending_spin.balance = float(balance_str)

                return

            if is_cascade and self._pending_spin is not None:
                tw: float = float(get("tw", 0))
                self._pending_spin.win = tw
                self._pending_spin.is_winning = tw > 0
                balance_str: str = get("balance", "0").replace(",", "")
                self._pending_spin.balance = float(balance_str)

                rmul: t.Optional[str] = params.get("rmul", [None])[0]
                if rmul:
                    multiplier: int = sum(int(x) for x in re.split(r'[~;]', rmul)) if rmul else 0
                    self._pending_spin.multiplier = max(self._pending_spin.multiplier, multiplier)
            else:
                if self._pending_spin is not None:
                    self._flush_pending_spin()

                self._spin_count += 1
                bet: float = float(get("c", 0)) * float(get("l", 0))
                rmul: t.Optional[str] = params.get("rmul", [None])[0]
                multiplier: int = sum(int(x) for x in re.split(r'[~;]', rmul)) if rmul else 0

                self._pending_spin = SpinRecord(
                    spin_number=self._spin_count,
                    index=int(get("index", 0)),
                    bet=bet,
                    win=float(get("tw", 0)),
                    balance=float(get("balance", "0").replace(",", "")),
                    ntp=float(get("ntp", 0)),
                    reel_set=int(get("reel_set", 0)),
                    has_tumble="tmb" in params,
                    is_winning=float(get("tw", 0)) > 0,
                    multiplier=multiplier,
                    event="",
                    free_spins="",
                    fs_win=0.0,
                )

                if self._spin_count >= self.spin_limit:
                    self._done.set()

        elif na == "c" and self._pending_spin is not None:
            fsend_total: t.Optional[str] = get("fsend_total")
            if fsend_total == "1":
                fswin_total: float = float(get("fswin_total", 0))
                self._pending_spin.fs_win = fswin_total
                self._pending_spin.win = fswin_total
                self._in_free_spins = False
                return

            tw: float = float(get("tw", 0))
            if tw > self._pending_spin.win:
                self._pending_spin.win = tw
                self._pending_spin.is_winning = tw > 0

            balance_str: str = get("balance", "0").replace(",", "")
            self._pending_spin.balance = float(balance_str)

        elif na == "s" and "s" not in params:
            if self._pending_spin is not None:
                balance_str: str = get("balance", "0").replace(",", "")
                self._pending_spin.balance = float(balance_str)
                self._flush_pending_spin()

    def _flush_pending_spin(self) -> None:
        spin: t.Optional[SpinRecord] = self._pending_spin
        self._pending_spin = None
        self.spins.append(spin)
        logger.info(
            f"[Spin {spin.spin_number}] bet={spin.bet} win={spin.win} balance={spin.balance} mul={spin.multiplier}"
        )

    async def wait_until_done(self) -> None:
        await self._done.wait()

        if self._pending_spin is not None:
            self._flush_pending_spin()

    def calculate_analytics(self) -> dict:
        spins: list[SpinRecord] = self.spins
        if not spins:
            return {}

        starting_balance: float = STARTING_BALANCE
        final_balance: float = spins[-1].balance
        balances: list[float] = [s.balance for s in spins]
        bets: list[float] = [s.bet for s in spins]
        wins: list[float] = [s.win for s in spins if s.is_winning]

        total_wagered: float = sum(bets)
        total_returned: float = sum(s.win for s in spins)

        longest_loss: int = 0
        longest_win: int = 0
        cur_loss: int = 0
        cur_win: int = 0
        for s in spins:
            if s.is_winning:
                cur_win += 1
                cur_loss = 0
            else:
                cur_loss += 1
                cur_win = 0
            longest_win = max(longest_win, cur_win)
            longest_loss = max(longest_loss, cur_loss)

        peak: float = starting_balance
        max_drawdown: float = 0.0
        peak_spin: int = 0
        for i, s in enumerate(spins):
            if s.balance > peak:
                peak = s.balance
                peak_spin = i + 1
            drawdown: float = peak - s.balance
            max_drawdown = max(max_drawdown, drawdown)

        downtrend_spin: t.Optional[int] = None
        running_max: float = balances[-1]
        for i in range(len(spins) - 1, -1, -1):
            if balances[i] >= running_max:
                running_max = balances[i]
                downtrend_spin = i + 1
            else:
                break

        ldw_count: int = sum(1 for s in spins if 0 < s.win < s.bet)

        from collections import Counter
        win_counts: Counter = Counter(s.win for s in spins if s.is_winning)
        most_frequent_win: float
        most_frequent_count: int
        most_frequent_win, most_frequent_count = win_counts.most_common(1)[0] if win_counts else (0, 0)
        free_spins_list: list[SpinRecord] = [s for s in spins if s.event == "Free Spins"]
        first_bonus_spin: t.Optional[int] = next((s.spin_number for s in spins if s.event), None)

        return {
            "Balance": {
                "Starting balance": starting_balance,
                "Final balance": final_balance,
            },
            "Wins": {
                "Total winning combinations": len(wins),
                "Largest win": max(wins) if wins else 0,
                "Smallest win": min(wins) if wins else 0,
                "Most frequent win amount": most_frequent_win,
                "Most frequent win count": most_frequent_count,
            },
            "Bonuses": {
                "Free Spins count": len(free_spins_list),
                "Bonus Game count": 0,
                "First bonus after spin": first_bonus_spin if first_bonus_spin else "N/A",
                "Max Free Spins bonus total": max((s.fs_win for s in free_spins_list), default="N/A"),
            },
            "Streaks": {
                "Longest losing streak": longest_loss,
                "Longest winning streak": longest_win,
            },
            "Balance dynamics": {
                "Peak balance": round(peak, 2),
                "Spin at peak balance": peak_spin,
                "Max drawdown": round(max_drawdown, 2),
                "Irreversible downtrend started at spin": downtrend_spin,
            },
            "Financial metrics": {
                "Net Result": round(final_balance - starting_balance, 2),
                "Total Wagered": round(total_wagered, 2),
                "Total Returned": round(total_returned, 2),
            },
            "Return": {
                "Observed Return (%)": round(total_returned / total_wagered * 100, 2) if total_wagered else 0,
                "Session Return (%)": round(final_balance / starting_balance * 100, 2),
            },
            "Hit Rate": {
                "Hit Rate (%)": round(len(wins) / len(spins) * 100, 2),
            },
            "LDW": {
                "LDW count": ldw_count,
                "LDW (%)": round(ldw_count / len(spins) * 100, 2),
            },
        }

    def save_xlsx(self, path: str = "spins.xlsx") -> None:
        wb: Workbook = Workbook()

        ws_spins = wb.active
        ws_spins.title = "Spins"

        headers: list[str] = ["Spin number", "Bet", "Win", "Balance", "Multiplier", "Event / Bonus", "Free Spins", "FS Win"]
        header_fill: PatternFill = PatternFill("solid", start_color="2F4F8F")
        header_font: Font = Font(bold=True, color="FFFFFF", name="Arial")

        for col, h in enumerate(headers, 1):
            cell = ws_spins.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for s in self.spins:
            ws_spins.append([
                s.spin_number, s.bet, s.win, s.balance,
                s.multiplier if s.multiplier else "",
                s.event or "", s.free_spins or "",
                s.fs_win if s.fs_win else "",
            ])

        for col in ws_spins.columns:
            ws_spins.column_dimensions[col[0].column_letter].width = 16

        ws_analytics = wb.create_sheet("Analytics")
        analytics: dict = self.calculate_analytics()

        section_fill: PatternFill = PatternFill("solid", start_color="2F4F8F")
        section_font: Font = Font(bold=True, color="FFFFFF", name="Arial")
        label_font: Font = Font(name="Arial")
        value_font: Font = Font(name="Arial")

        row: int = 1
        for section, metrics in analytics.items():
            cell = ws_analytics.cell(row=row, column=1, value=section)
            cell.fill = section_fill
            cell.font = section_font
            ws_analytics.cell(row=row, column=2).fill = section_fill
            row += 1

            for label, value in metrics.items():
                ws_analytics.cell(row=row, column=1, value=label).font = label_font
                ws_analytics.cell(row=row, column=2, value=value).font = value_font
                row += 1

            row += 1

        ws_analytics.column_dimensions["A"].width = 42
        ws_analytics.column_dimensions["B"].width = 20

        wb.save(path)
        logger.info(f"Saved: {path}")
