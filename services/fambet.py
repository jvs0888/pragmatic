import asyncio
import functools
import os
import typing as t
from dotenv import load_dotenv
from pathlib import Path

from playwright.async_api import Playwright, async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from loggers.logger import logger
    from base.playwright_async import PlaywrightAsync
    from services.spin_manager import SpinCollector
    from services.provider_manager import ProviderManager
except ImportError as ie:
    exit(f"{ie} :: {Path(__file__).resolve()}")

load_dotenv()

def playwright_initiator(func):
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        playwright_context: Playwright = await async_playwright().start()
        try:
            kwargs['playwright'] = playwright_context
            return await func(*args, **kwargs)
        except Exception as e:
            logger.exception(f'exception in "{func.__name__}" => {e}')
            return False
        finally:
            await playwright_context.stop()
    return wrapper


class Fambet(PlaywrightAsync):
    def __init__(self) -> None:
        super().__init__()
        self.base_url: str = "https://fambet.com"
        self.start_url: str = f"{self.base_url}/ca/"
        self.games_num: int = 7
        self.games_titles: list = ["Top games in Canada", "New"]
        self.games: dict = dict()

    async def login(self, email: str, password: str) -> None:
        login_button = await self.wait_for_selector(
            selector="//stb-high-custom-button[@data-testid='headerLoginBtnDesktop']"
        )

        if not login_button:
            logger.error("Login failed")
            return

        await login_button.click()

        email_input = await self.wait_for_selector(selector="//input[@data-testid='email']")
        if email_input:
            await self.human_type(email_input, email)

        password_input = await self.wait_for_selector(selector="//input[@data-testid='password']")
        await self.human_type(password_input, password)

        submit_button = await self.wait_for_selector(selector="//stb-high-custom-button[@data-testid='btnLogin']")
        await submit_button.click()

    async def find_games(self) -> dict:
        game_lists = await self.page.query_selector_all("//stb-game-list[@class='games-list']")

        for game_list in game_lists:
            link = await game_list.query_selector("stb-games-slider-title a")
            if not link:
                continue

            text = (await link.inner_text()).strip()

            if text.lower() in [t.lower() for t in self.games_titles]:
                await game_list.scroll_into_view_if_needed()
                games = await game_list.query_selector_all("[class='grid-games__item']")

                games_data: list = []
                for game in games[:self.games_num]:
                    game_link = await game.query_selector("[class='thumbnail-game-link']")

                    if not game_link:
                        continue

                    name = await game_link.get_attribute("aria-label")
                    url = await game_link.get_attribute("href")
                    games_data.append({"name": name, "url": url})

                self.games[text] = games_data

    async def process_games(self) -> None:
        provider_manager: ProviderManager = ProviderManager()

        for category, games in self.games.items():
            logger.info(f"Processing category: {category} | Games count: {len(games)}")

            for game in games:
                new_page: t.Optional = None
                provider_name: t.Optional[str] = None
                try:
                    new_page = await self.context.new_page()
                    logger.info(f"Opening game: {game['name']}")

                    url: str = self.base_url + game["url"]
                    await new_page.goto(url, wait_until="domcontentloaded")
                    await asyncio.sleep(2)

                    provider_selector = await new_page.wait_for_selector(
                        selector="//p[@class='game-header__provider']",
                        timeout=10000
                    )
                    provider_name = (await provider_selector.inner_text()).strip()

                    if provider_name == 'Novomatic':
                        currency_alert = await new_page.wait_for_selector(
                            selector="//stb-high-custom-button[@stbclosedialog]"
                        )
                        if currency_alert:
                            await currency_alert.click()


                    logger.info(f"Game: {game['name']} | Provider: {provider_name}")

                    game["provider"] = provider_name

                    rtp_data: t.Optional[dict] = await provider_manager.process_game(
                        provider_name=provider_name,
                        page=new_page,
                        game_name=game["name"]
                    )

                    if rtp_data and "rtp" in rtp_data:
                        game["rtp"] = rtp_data["rtp"]
                        logger.info(f"RTP saved: {game['name']} | RTP: {game['rtp']}")
                    else:
                        game["rtp"] = None
                        logger.warning(f"RTP not found for game: {game['name']}")

                except Exception as e:
                    logger.error(f"Error processing game {game.get('name', 'Unknown')}: {e}")
                    game["provider"] = None
                    game["rtp"] = None
                finally:
                    if new_page:
                        try:
                            await new_page.close()
                        except Exception as e:
                            logger.error(f"Error closing page: {e}")

                if provider_name == 'Novomatic':
                    currency_alert = await self.wait_for_selector(
                        selector="//stb-high-custom-button[@stbclosedialog]"
                    )
                    if currency_alert:
                        await currency_alert.click()

                await asyncio.sleep(3)

    def save_xlsx(self, path: str = "fambet_rtp.xlsx") -> None:
        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = "RTP Data"

        headers: list = ["Slot name", "Provider", "Real RTP"]
        header_fill: PatternFill = PatternFill("solid", start_color="2F4F8F")
        header_font: Font = Font(bold=True, color="FFFFFF", name="Arial")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row: int = 2
        for category, games in self.games.items():
            for game in games:
                ws.cell(row=row, column=1, value=game.get("name", "N/A"))
                ws.cell(row=row, column=2, value=game.get("provider", "N/A"))

                rtp = game.get("rtp")
                if rtp is not None:
                    ws.cell(row=row, column=3, value=str(rtp))
                else:
                    ws.cell(row=row, column=3, value="N/A")

                row += 1

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        wb.save(path)
        logger.info(f"RTP data saved: {path}")

    @playwright_initiator
    async def execute(self, playwright: Playwright):
        try:
            if not await self.init_browser_process():
                logger.error("Browser process not initialized")
                return

            await asyncio.sleep(5)

            if not await self.connect_cdp_session(playwright):
                logger.error("Connection not initialized")
                return

            await self.page.goto(self.start_url)
            await self.login(
                email=os.getenv("EMAIL"),
                password=os.getenv("PASSWORD")
            )

            alert_message = await self.wait_for_selector(
                selector="//button[@data-testid='btnCloseNewBonus']",
                timeout=15000
            )
            if alert_message:
                await alert_message.click()

            await self.find_games()
            await asyncio.sleep(3)

            await self.process_games()

            self.save_xlsx()
            logger.info("All games processed successfully!")
        except Exception as e:
            logger.exception(e)
        finally:
            await self.close()

    def run(self):
        try:
            asyncio.run(self.execute())
        except (KeyboardInterrupt, asyncio.CancelledError):
            raise
        except Exception:
            pass
